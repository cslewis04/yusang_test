### STEP0. 라이브러리 가져오기
import requests
import zipfile
import os
import pandas as pd
import numpy as np
import re
import math
import pickle
import warnings
import OpenDartReader
import time
import json
import streamlit as st
from streamlit_lottie import st_lottie

from bs4 import BeautifulSoup
from io import BytesIO
from datetime import datetime, timedelta, date
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, PatternFill, colors, Color, Font

warnings.filterwarnings(action='ignore')
API_KEY = '9ed5bc3d6d1fee4f927c5d6a44eb5368d222824c'
dart = OpenDartReader(API_KEY)


### STEP1. 증권신고서(지분증권)및 증권발행실적보고서 정보 가져오기
def get_rcept_no(bgn_de, end_de):
    info_df = dart.list(start=bgn_de, end=end_de, kind_detail='C001', final=False)
    info_df = info_df.loc[info_df.corp_cls.isin(['Y', 'K'])]
    info_df = info_df.loc[info_df.report_nm.str.contains('증권신고서') | info_df.report_nm.str.contains('증권발행실적보고서')]
    info_df['구분'] = np.where(info_df.report_nm.str.contains('증권신고서'), '1', '2')
    info_df['비고'] = np.where(info_df.report_nm.str.contains('기재정정'), '정정',
                             np.where(info_df.report_nm.str.contains('발행조건확정'), '발행', '최초'))
    rcept_info = list(
        info_df['rcept_no'] + '_' + info_df['corp_cls'] + info_df['구분'] + '_' + info_df['corp_name'] + '_' + info_df[
            '비고'])

    print('보고서수 : ', len(rcept_info))
    st.write('보고서 수(신규상장 및 이전상장 제외 전) : ', len(rcept_info))
    return (rcept_info)


### STEP2. 테이블에서 값 가져오기
def get_table(xml_text, title, gubun):
    try:
        # 표 제목을 가진 테이블 찾기(포함 유무로 판단, 빈칸 고려)
        table_src = re.findall('{}.*?</TABLE-GROUP>'.format(title), xml_text, re.DOTALL)[0].replace('TE', 'TD').replace(
            'TU', 'TD')
        table = pd.read_html(table_src)
        df = table[1] if len(table) > 1 and len(table[0]) == 1 else table[0]

        # 표에 칼럼명이 숫자로 임의로 지정되었을 때 첫번째 행의 값을 칼럼명으로 지정하기
        if all(isinstance(col, int) for col in df.columns):
            df.columns = df.iloc[0]

        # 정정인 경우 정정 후 테이블 찾기
        if gubun == '정정':
            num = min(10, len(table))
            for i in range(2, num):
                if len(table[i].columns) == len(df.columns):
                    if all(isinstance(col, int) for col in table[i].columns):
                        table[i].columns = table[i].iloc[0]
                    if all(table[i].columns == df.columns) == True:
                        df = table[i]
        return df
    except Exception:
        return pd.DataFrame()


def get_value_from_table(df, find, col):
    """
    :df : 가져올 테이블
    :find : 표에서 찾고 싶은 값 (빈칸 및 괄호내용 제거, 포함유무로 확인)
    :col : 표에서 추출될 칼럼명 또는 칼럼위치 (빈칸 제거)
    """
    try:
        # 테이블 내 값에 모든 빈칸을 제거하기
        df = df.applymap(lambda x: x.replace(' ', '') if isinstance(x, str) else x)

        # 찾고 싶은 값을 포함하는 행 찾기
        find_row = df[df.apply(lambda x: x.astype(str).str.contains(find)).any(axis=1)].index[0]

        if isinstance(df.columns, pd.MultiIndex):
            # 칼럼명에 빈칸 제거
            df.columns = df.columns.map(lambda x: (x[0].replace(' ', ''), x[1].replace(' ', '')))
        else:
            # 칼럼명에 빈칸 및 괄호내용 제거
            df.columns = df.columns.str.replace('\(.*\)', '', regex=True).str.replace(' ', '')

        # 찾은 행으로부터 해당 칼럼값 추출
        if isinstance(col, int):
            value = df.iloc[find_row, col]
        else:
            value = df.loc[find_row, col]
        return value
    except:
        return '추출불가'


def get_value_from_id(xml_text, table_id, value_id):
    table = re.findall('<TABLE-GROUP ACLASS="' + table_id + '".*?</TABLE-GROUP>', xml_text, re.DOTALL)[0].replace('TE',
                                                                                                                  'TD').replace(
        'TU', 'TD')
    num = len(re.findall('"' + value_id + '".*?</TD>', table, re.DOTALL))
    if num > 1:
        value = []
        for i in range(num):
            value.append(
                re.findall('"' + value_id + '".*?</TD>', table, re.DOTALL)[i].replace('><', '').split('>')[1].split(
                    '<')[0])
    else:
        value = \
        re.findall('"' + value_id + '".*?</TD>', table, re.DOTALL)[0].replace('><', '').split('>')[1].split('<')[0]
    return value


### STEP3. 증권신고서와 증권발행실적보고서 추출하기
def get_singo_balh(rcept_no):
    xml_text = dart.document(rcept_no)
    xml_text = xml_text.replace("\n", "").replace("모집(매출) 방법", "모집(매출)방법")

    soup = BeautifulSoup(xml_text, 'html.parser')
    table_src = [str(x) for x in soup.find_all('table')]
    tables = [pd.read_html(x)[0] for x in table_src]
    try:
        f_dfs = [x for x in tables if "모집(매출)방법" in str(x)][0]

        # 표에 칼럼명이 숫자로 임의로 지정되었을 때 첫번째 행의 값을 칼럼명으로 지정하기
        if all(isinstance(col, int) for col in f_dfs.columns):
            f_dfs.columns = f_dfs.iloc[0]
        value = f_dfs['모집(매출)방법'][len(f_dfs) - 1]
        return value
    except:
        return '확인필요'


def get_singo(info):
    xml_text = dart.document(info[:14])
    xml_text = xml_text.replace("\n", "")
    gubun = info[-2:]
    rows = []

    if get_value_from_id(xml_text, 'ACC_ES', 'PSSRP_KND3') not in ['신규상장', '이전상장']:
        try:
            iljeong_old = get_table(xml_text, '공모일정 등에 관한 사항', gubun)  # 1.공모개요
            iljeong_remove = iljeong_old[~iljeong_old.iloc[:, 1].str.contains('공고|폐지|무상증자')]
            iljeong = iljeong_remove.iloc[::-1].reset_index()
            iljeong = iljeong.iloc[:, 1:]
            sanchul = get_table(xml_text, '배정비율 산출', gubun)  # 2.공모방법
            sanjeong = get_table(xml_text, '산정표', gubun)  # 1.공모개요
            gongmo = get_table(xml_text, '공모방법', gubun)  # 2.공모방법

            jeungja_mth = get_value_from_id(xml_text, 'PBO', 'PB_MTH')
            isa_dt = get_value_from_table(iljeong, '이사회결의', 0) if not iljeong.empty else '테이블X'
            baej_dt = get_value_from_id(xml_text, 'SCHD', 'ASN_BAS_DT')
            sinju_sangj_dt = get_value_from_table(iljeong, '신주인수권증서상장', 0) if not iljeong.empty else '테이블X'
            guju_dt = get_value_from_id(xml_text, 'SCHD', 'SSC_PRD')
            ilban_dt = get_value_from_table(iljeong, '일반공모청약', 0) if not iljeong.empty else '테이블X'
            nabib_dt = get_value_from_id(xml_text, 'SCHD', 'PYM_PRD')
            sanj_dt = get_value_from_table(iljeong, '상장예정일', 0) if not iljeong.empty else '테이블X'
            exp_price = get_value_from_id(xml_text, 'PBO', 'PB_VAL')
            done_cnt = get_value_from_table(sanchul, 'C.발행주식총수', 1) if not sanchul.empty else '테이블X'
            exp_cnt = get_value_from_id(xml_text, 'PBO', 'STK_CNT')
            exp_all_price = get_value_from_id(xml_text, 'PBO', 'PB_TOT')
            discnt_per = get_value_from_table(sanjeong, '할인율', '거래량') if not sanjeong.empty else '테이블X'
            jeungja_per = get_value_from_table(sanchul, '증자비율', 1) if not sanchul.empty else '테이블X'
            daepyo_yn = get_value_from_id(xml_text, 'ACC', 'ACC_KND')
            insu_nm = get_value_from_id(xml_text, 'ACC', 'ACC_NMT')
            insu_mth = get_value_from_id(xml_text, 'ACC', 'ACC_MTH')
            condition = get_value_from_id(xml_text, 'ACC', 'ACC_PRI')
            uli_saju = get_value_from_table(gongmo, '우리사주', 1) if not gongmo.empty else '테이블X'

            row = {'증자방법': jeungja_mth, '이사회결의일': isa_dt, '신주배정기준일': baej_dt, '신주인수권상장일': sinju_sangj_dt,
                   '구주주청약일': guju_dt, '일반청약일': ilban_dt, '납입일': nabib_dt, '상장일': sanj_dt, '예정발행가': exp_price,
                   '기발행주식수': done_cnt, '예정발행주식수': exp_cnt, '예정발행금액': exp_all_price,
                   '할인율': discnt_per, '증자비율': jeungja_per, '우리사주': uli_saju,
                   '대표유무': daepyo_yn, '인수방식': insu_mth, '인수단': insu_nm, '특이조건': condition}
            rows.append(row)
        except Exception as e:
            print(info[:14] + '_Error!_' + str(e))
    return rows


def get_silj(info):
    xml_text = dart.document(info[:14])
    xml_text = xml_text.replace("\n", "")
    gubun = info[-2:]
    rows = []

    try:
        hyeonh = get_table(xml_text, '청약 및 배정현황', gubun)
        insu = get_table(xml_text, '인수기관별 인수금액', gubun)
        baej_cnt = get_value_from_table(hyeonh, '일반', ('최종배정현황', '수량'))
        cheongy_cnt = get_value_from_table(hyeonh, '일반', ('청약현황', '수량'))
        gongm_cnt = get_value_from_table(hyeonh, '계', ('최종배정현황', '수량'))
        gongm_price = get_value_from_table(hyeonh, '계', ('최종배정현황', '금액'))
        third_cnt = get_value_from_table(hyeonh, '3자', ('최종배정현황', '수량'))
        third_price = get_value_from_table(hyeonh, '3자', ('최종배정현황', '금액'))
        third_per = get_value_from_table(hyeonh, '3자', ('최종배정현황', '비율'))
        insu_nm = get_value_from_table(insu, '증권', 0) if not insu.empty else ''
        jugwansa_cnt = get_value_from_table(insu, '증권', 1) if not insu.empty else 0
        jugwansa_price = get_value_from_table(insu, '증권', 2) if not insu.empty else 0
        final_cnt = get_value_from_table(hyeonh, '계', ('최종배정현황', '수량'))
        final_price = get_value_from_table(hyeonh, '계', ('최종배정현황', '금액'))

        row = {'일반배정주식수': baej_cnt, '일반청약주식수': cheongy_cnt, '공모배정주식수': gongm_cnt, '공모배정금액': gongm_price,
               '3자배정청약주식수': third_cnt, '3자배정청약금액': third_price, '3자배정배정율': third_per,
               '인수기관': insu_nm, '주관사인수주식수': jugwansa_cnt, '주관사인수금액': jugwansa_price,
               '최종발행주식수': final_cnt, '최종발행금액': final_price}
        rows.append(row)
    except Exception as e:
        print(info[:14] + '_Error!_' + str(e))
    return rows


### STEP4. 값을 원하는 형식으로 처리하기
def to_date(var):
    if var == '추출불가':
        value = ''
    else:
        numbers = ''.join(filter(str.isdigit, var))
        if len(numbers) > 8:
            value = datetime.strptime(numbers[:8], '%Y%m%d').date()
        elif len(numbers) == 8:
            value = datetime.strptime(numbers, '%Y%m%d').date()
        else:
            value = ''
    return value


def to_int(var):
    gwalho = re.sub(r'\([^)]*\)', '', str(var))
    sosu = gwalho.split(".")[0]
    numbers = re.sub(r'\D', '', sosu)
    if len(numbers) == 0:
        value = 0
    else:
        value = int(numbers)
    return value


def to_per(var):
    if var == '추출불가':
        value = '추출불가'
    elif '%' in str(var):
        value = float(re.sub(r'[^0-9.]', '', var)) / 100
    else:
        value = '확인필요'
    return value


def to_per_gwalho(var):
    if var == '추출불가':
        value = '추출불가'
    elif '(' in str(var):
        var = re.findall(r'\((.*?)\)', var)[0]
        value = float(re.sub(r'[^0-9.]', '', var)) / 100
    elif var == '-':
        value = 0
    else:
        value = '확인필요'
    return value


def to_short(var):
    replace_dict = {
        '(주)': '',
        '증권': '',
        '투자': '',
        '금융': '',
        '에셋': '',
        '아이비케이': 'IBK',
        '케이비': 'KB'
    }
    for key, value in replace_dict.items():
        var = var.replace(key, value)
    return var


### STEP5. 원하는 보고서 형식으로 내보내기¶
def get_report(info):
    rows1 = []
    rows2 = []
    for i in reversed(range(len(info))):
        corp_nm = info[i][18:-3]
        mkt_type = '유' if info[i][15:16] == 'Y' else '코'
        rcept_no = 'https://dart.fss.or.kr/dsaf001/main.do?rcpNo=' + info[i][:14]
        if info[i][16:17] == '1':
            try:
                if info[i][-2:] in ['최초', '정정'] and get_singo(info[i]) != []:
                    data = get_singo(info[i])[0]
                    jeungja_mth = data['증자방법']
                    isa_dt = to_date(data['이사회결의일'])
                    baej_dt = to_date(data['신주배정기준일'])
                    sinju_sangj_dt = to_date(data['신주인수권상장일'])
                    guju_dt = to_date(data['구주주청약일'])
                    ilban_dt = to_date(data['일반청약일'])
                    nabib_dt = to_date(data['납입일'])
                    sanj_dt = to_date(data['상장일'])
                    if info[i][-2:] == '최초':
                        exp_price = to_int(data['예정발행가'])
                        done_cnt = to_int(data['기발행주식수'])
                        exp_cnt = to_int(data['예정발행주식수'])
                        exp_all_price = to_int(data['예정발행금액'])
                        discnt_per = to_per(data['할인율'])
                        jeungja_per = exp_cnt / done_cnt if done_cnt != 0 else '확인필요'
                        uli_saju_per = 0 if to_per_gwalho(data['우리사주']) == '추출불가' else to_per_gwalho(data['우리사주'])
                        insu_mth = '모집주선' if data['대표유무'] == '주선' else (
                            data['인수방식'][0] if type(data['인수방식']) == list else data['인수방식'])
                        sangj_yn = '비상장' if sinju_sangj_dt == '' else '상장'

                        num = len(data['대표유무']) if type(data['대표유무']) == list else 1
                        daepyo = []
                        for j in range(num):
                            if type(data['대표유무']) == list:
                                if data['대표유무'][j] in ['대표', '공동', '주선']:
                                    daepyo.append(to_short(data['인수단'][j]))
                            else:
                                if data['대표유무'] in ['대표', '공동', '주선']:
                                    daepyo.append(to_short(data['인수단']))
                        daepyo_nm = ', '.join(daepyo)
                        for j in range(num):
                            if type(data['대표유무']) == list:
                                insu_nm = to_short(data['인수단'][j])
                                condition = data['특이조건'][j]
                            else:
                                insu_nm = to_short(data['인수단'])
                                condition = data['특이조건']
                            row = {'비고': '', '보고서주소': rcept_no, '법인명': corp_nm, '시장구분': mkt_type,
                                   '증자방법': jeungja_mth, '이사회결의일': isa_dt, '신주배정기준일': baej_dt,
                                   '신주인수권상장일': sinju_sangj_dt, '구주주청약초일': guju_dt, '일반청약초일': ilban_dt,
                                   '납입일': nabib_dt, '상장일': sanj_dt, '예정발행가': exp_price, '기발행주식수': done_cnt,
                                   '예정발행주식수': exp_cnt, '예정발행금액': exp_all_price, '할인율': discnt_per, '증자비율': jeungja_per,
                                   '우리사주배정비율': uli_saju_per, '대표주관': daepyo_nm, '인수방식': insu_mth, '인수단': insu_nm,
                                   '정액수수료': '', '기본수수료율': '', '추가수수료율': '', '특이조건': condition, '인수모집비율': '',
                                   '인수모집금액': '', '수수료금액': '', '1차발행가': '', '2차발행가': '',
                                   '확정발행가': '', '발행가증감율': '', '신주인수권상장여부': sangj_yn}
                            rows1.append(row)
                            st.write('<p style="font-size:14px; color:black">' + '- 문서 ' + info[i][:14] + ' 추출 완료!</p>',
                                     unsafe_allow_html=True)
                    elif info[i][-2:] == '정정':
                        row = {'비고': info[i][-2:], '보고서주소': rcept_no, '법인명': corp_nm, '시장구분': mkt_type,
                               '증자방법': jeungja_mth,
                               '이사회결의일': isa_dt, '신주배정기준일': baej_dt, '신주인수권상장일': sinju_sangj_dt,
                               '구주주청약초일': guju_dt, '일반청약초일': ilban_dt, '납입일': nabib_dt, '상장일': sanj_dt,
                               '예정발행가': '', '기발행주식수': '', '예정발행주식수': '', '예정발행금액': '', '할인율': '', '증자비율': '',
                               '우리사주배정비율': '', '대표주관': '', '인수방식': '', '인수단': '', '정액수수료': '', '기본수수료율': '',
                               '추가수수료율': '', '특이조건': '', '인수모집비율': '', '인수모집금액': '', '수수료금액': '',
                               '1차발행가': '', '2차발행가': '', '확정발행가': '', '발행가증감율': '', '신주인수권상장여부': ''}
                        rows1.append(row)
                        st.write('<p style="font-size:14px; color:black">' + '- 문서 ' + info[i][:14] + ' 추출 완료!</p>',
                                 unsafe_allow_html=True)
                elif info[i][-2:] == '발행':
                    row = {'비고': info[i][-2:] + '(' + get_singo_balh(info[i][:14]) + ')', '보고서주소': rcept_no,
                           '법인명': corp_nm,
                           '시장구분': mkt_type, '증자방법': '', '이사회결의일': '', '신주배정기준일': '', '신주인수권상장일': '',
                           '구주주청약초일': '', '일반청약초일': '', '납입일': '', '상장일': '', '예정발행가': '', '기발행주식수': '',
                           '예정발행주식수': '', '예정발행금액': '', '할인율': '', '증자비율': '', '우리사주배정비율': '',
                           '대표주관': '', '인수방식': '', '인수단': '', '정액수수료': '', '기본수수료율': '', '추가수수료율': '',
                           '특이조건': '', '인수모집비율': '', '인수모집금액': '', '수수료금액': '',
                           '1차발행가': '', '2차발행가': '', '확정발행가': '', '발행가증감율': '', '신주인수권상장여부': ''}
                    rows1.append(row)
                    st.write('<p style="font-size:14px; color:black">' + '- 문서 ' + info[i][:14] + ' 추출 완료!</p>',
                             unsafe_allow_html=True)
            except Exception as e:
                print(info[i] + '_Error!_' + str(e))
        elif info[i][16:17] == '2' and info[i][-2:] == '최초':
            try:
                data = get_silj(info[i])[0]
                baej_cnt = to_int(data['일반배정주식수'])
                cheongy_cnt = to_int(data['일반청약주식수'])
                competi_per = cheongy_cnt / baej_cnt if baej_cnt != 0 else 0
                gongm_cnt = to_int(data['공모배정주식수'])
                gongm_price = to_int(data['공모배정금액'])
                third_cnt = to_int(data['3자배정청약주식수'])
                third_price = to_int(data['3자배정청약금액'])
                third_per = to_int(data['3자배정배정율'])
                insu_nm = to_short(data['인수기관']) if data['인수기관'] != '추출불가' else ''
                jugwansa_cnt = to_int(data['주관사인수주식수'])
                jugwansa_price = to_int(data['주관사인수금액'])
                final_cnt = to_int(data['최종발행주식수'])
                final_price = to_int(data['최종발행금액'])
                jugwansa_per = jugwansa_cnt / final_cnt if final_cnt != 0 else '확인필요'

                row = {'비고': '', '보고서주소': rcept_no, '법인명': corp_nm, '시장구분': mkt_type,
                       '일반배정주식수': baej_cnt, '일반청약주식수': cheongy_cnt, '일반경쟁률': competi_per,
                       '공모배정주식수': gongm_cnt, '공모배정금액': gongm_price, '공모배정율': '',
                       '3자배정청약주식수': third_cnt, '3자배정청약금액': third_price, '3자배정배정율': third_per,
                       '주관사인수주식수': jugwansa_cnt, '주관사인수금액': jugwansa_price, '주관사인수율': jugwansa_per,
                       '최종발행주식수': final_cnt, '최종발행금액': final_price}
                rows2.append(row)
                st.write('<p style="font-size:14px; color:black">' + '- 문서 ' + info[i][:14] + ' 추출 완료!</p>',
                         unsafe_allow_html=True)
            except Exception as e:
                print(info[i] + '_Error!_' + str(e))
        else:
            row = {'비고': info[i][-2:], '보고서주소': rcept_no, '법인명': corp_nm, '시장구분': mkt_type,
                   '일반배정주식수': '', '일반청약주식수': '', '일반경쟁률': '', '공모배정주식수': '', '공모배정금액': '', '공모배정율': '',
                   '3자배정청약주식수': '', '3자배정청약금액': '', '3자배정배정율': '',
                   '주관사인수주식수': '', '주관사인수금액': '', '주관사인수율': '', '최종발행주식수': '', '최종발행금액': ''}
            rows2.append(row)
            st.write('<p style="font-size:14px; color:black">' + '- 문서 ' + info[i][:14] + ' 추출 완료!</p>',
                     unsafe_allow_html=True)
    result1 = pd.DataFrame(rows1)
    result2 = pd.DataFrame(rows2)
    return (result1, result2)


### STEP5. 웹페이지 레이아웃 및 엑셀 형식 설정하기
# 애니메이션 및 보고서 제목 삽입
def load_lottie():
    with open('./resources/report.json', 'r', encoding='utf-8-sig') as st_json:
        return json.load(st_json)


empty1, col1, col2 = st.columns([0.05, 0.3, 0.8])
with empty1:
    st.empty()
with col1:
    lottie = load_lottie()
    st_lottie(lottie, speed=1, loop=True, width=150, height=150, )
with col2:
    ''
    ''
    st.title('유상증자 집계현황')

# 날짜 선택
start_date = st.date_input('시작일', value=date.today(), max_value=date.today())
max_date = min(start_date + timedelta(days=31), date.today())
end_date = st.date_input('종료일', value=start_date, min_value=start_date, max_value=max_date)

# 조회 및 다운 버튼 생성
if st.button("조회"):
    bgn_de = datetime.strftime(start_date, '%Y%m%d')
    end_de = datetime.strftime(end_date, '%Y%m%d')
    info = get_rcept_no(bgn_de, end_de)
    result1, result2 = get_report(info)

    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(result1, index=False, header=True):
        ws.append(r)
    ws.append([])
    for r in dataframe_to_rows(result2, index=False, header=True):
        ws.append(r)
    for column_cells in ws.columns:
        for cell in ws[column_cells[0].column_letter]:
            cell.font = Font(size=9)
            if column_cells[0].column_letter == 'B' and cell.row not in [1, len(result1) + 2, len(result1) + 3]:
                cell.value = '=HYPERLINK("{}", "{}")'.format(cell.value, '링크')
            if cell.row == len(result1) + 2:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws.freeze_panes = "D2"

    wb.save('유상증자결과_' + bgn_de + '_' + end_de + '.xlsx')
    with open('유상증자결과_' + bgn_de + '_' + end_de + '.xlsx', 'rb') as f:
        data = f.read()
        st.download_button(label='다운', data=data, file_name='유상증자결과_' + bgn_de + '_' + end_de + '.xlsx',
                           mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')